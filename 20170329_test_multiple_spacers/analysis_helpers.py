#!/usr/bin/env python3

import color_me
import itertools
import numpy as np
from collections import namedtuple
from pprint import pprint

SENSORS = [
        'on',
        'off',
        'rxb 11',
        'mhf 30',
        'mhf 37',
]
COLORS = {
        'on': color_me.ucsf.light_grey[0],
        'off': color_me.ucsf.light_grey[0],
        'rxb 11': color_me.ucsf.red[0],
        'mhf 30': color_me.ucsf.blue[0],
        'mhf 37': color_me.ucsf.blue[0],
}


class CleavageData:

    # Despite its name, '20170512_test_d1_d4.xlsx' contains data for all the 
    # spacers I've ever tested.
    def __init__(self, path='20170512_test_d1_d4.xlsx'):
        self.data = load_cleavage_data_from_xlsx(path)
        self.sensors, self.spacers = find_sensors_and_spacers(self.data)

    def calc_percent_cut(self, sensor, spacer, ligand):
        return calc_percent_cut(self.data, sensor, spacer, ligand)

    def calc_percent_change(self, sensor, spacer):
        return calc_percent_change(self.data, sensor, spacer)



def load_cleavage_data_from_xlsx(path):
    from openpyxl import load_workbook
    book = load_workbook('20170512_test_d1_d4.xlsx')
    sheet = book['Data']
    spacer = None
    sensor = None
    ligand = None
    band = None

    data = {}
    SPACER, DESIGN, LIGAND, BAND = 1, 2, 3, 4
    PIXELS_REP_1, PIXELS_REP_2, PIXELS_REP_3 = 5, 6, 7
    Index = namedtuple('Index', ['sensor', 'spacer', 'ligand', 'band'])

    row = 2  # Skip the header (indexing is from 1).
    is_row_empty = lambda i: sheet.cell(row=i, column=BAND).value is None

    while not is_row_empty(row):
        spacer = sheet.cell(row=row, column=SPACER).value or spacer
        sensor = sheet.cell(row=row, column=DESIGN).value or sensor
        ligand = sheet.cell(row=row, column=LIGAND).value or ligand
        band   = sheet.cell(row=row, column=BAND).value or band

        index  = Index(sensor, spacer, (ligand == '+'), int(band[0]))
        pixels = [sheet.cell(row=row, column=j).value for j in [5, 6, 7]]
        pixels = np.array([x for x in pixels if x is not None])

        if pixels.size:
            data[index] = pixels

        row += 1

    return data

def find_sensors_and_spacers(data, expected_sensors=SENSORS):
    sensors = {x.sensor for x in data}
    spacers = {x.spacer for x in data}

    if expected_sensors:
        unexpected_sensors = set(sensors) - set(expected_sensors)
        if unexpected_sensors:
            raise ValueError(f"found unexpected spacer(s): {unexpected_sensors}")

    return sensors, spacers

def calc_percent_cut(data, sensor, spacer, ligand):
    cut_px = data[sensor, spacer, ligand, 2]
    uncut_px = data[sensor, spacer, ligand, 4]
    return cut_px / (cut_px + uncut_px)

def calc_percent_change(data, sensor, spacer):
    apo = calc_percent_cut(data, sensor, spacer, False)
    holo = calc_percent_cut(data, sensor, spacer, True)
    return holo - apo

def unclump_points(y, yerr, x=0, dx=1):
    """
    Return x values such that the given points can be plotted (with error bars) 
    without any overlaps.  
    """

    # Determine which points are clashing with each other.
    n = len(y)
    clashes = np.zeros((n,n))

    for i, j in itertools.combinations(range(n), 2):
        i_min = y[i] - yerr[i]
        i_max = y[i] + yerr[i]
        j_min = y[j] - yerr[j]
        j_max = y[j] + yerr[j]

        clashes[i,j] = clashes[j,i] = \
                (i_min <= j_min <= i_max) or (i_min <= j_max <= i_max) or \
                (j_min <= i_min <= j_max) or (j_min <= i_max <= j_max)

    # Create groups of non-clashing points using a greedy algorithm.
    labels = {}
    next_label = 0
    sorted_indices = list(np.argsort(y))
    unlabeled_indices = lambda: (
            i for i in sorted_indices if i not in labels)

    for i in unlabeled_indices():
        group = [i]

        for j in unlabeled_indices():
            for ii in group:
                if clashes[j,ii]: break
            else:
                group.append(j)

        for ii in group:
            labels[ii] = next_label

        next_label += 1

    # Calculate an x-coordinate for each group.
    def label_to_offset(ii): #
        sign = 1 if ii % 2 else -1
        factor = (ii + 1) // 2
        return x + sign * factor * dx

    return np.array([label_to_offset(labels[i]) for i in range(n)])


