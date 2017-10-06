#!/usr/bin/env python3

"""
Usage:
    make_heatmap.py
"""

import re
import numpy as np
import pandas as pd
from pathlib import Path
from pprint import pprint

import docopt
import shutil
import matplotlib.pyplot as plt
import matplotlib as mpl

pd.options.display.width = shutil.get_terminal_size().columns

DISPLAY = ['spacer', 'design', 'ligand', 'band', 'pixels']

def parse_xlsx_dir(dir):
    dir = Path(dir)
    dfs = [parse_xlsx(p) for p in sorted(dir.glob('*.xlsx'))]
    df = pd.concat(dfs, ignore_index=True)

    # Return a sorted data frame, just to make it nicer to look at.  Use 
    # 'natsort' to handle things like 'd2' vs 'd10'.

    return natsort_spacers(df)

def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
        book = load_workbook(path)
        sheet = book['data']

        # Start by parsing the column headers.  Not all the spreadsheets will 
        # have the same data in the same columns, so we'll need this 
        # information when we're parsing the rows.

        headers = ['expt']
        cols = []
        is_col_empty = lambda j: sheet.cell(row=1, column=j).value is None
        j = 1  # indexing is from 1.

        while not is_col_empty(j):
            title = sheet.cell(row=1, column=j).value

            # Attempt to remove weird characters from the column titles, 
            # including parenthesized unit labels and punctuation.  This is 
            # mostly aesthetic, but it allows downstream code to use pull data 
            # out of the data frame using the nicer attribute syntax.

            slug = title.lower()
            slug = re.sub('\(.*\)', '', slug)
            slug = re.sub('\W', '', slug)

            ignore = 'cleaved', 'change', 'notes'
            if slug not in ignore:
                headers.append(slug)
                cols.append(j)

            j += 1

        # Parse the data from each row.  For certain columns ('expt', 'spacer', 
        # 'design', and 'ligand'), fill in missing values from previous rows.
        
        rows = []
        previous_row = []
        is_row_empty = lambda i: not any([
            sheet.cell(row=i, column=j).value for j in cols])
        i = 2  # Skip the header (indexing is from 1).

        def parse_cell(i, j, j0): #
            x = sheet.cell(row=i, column=j).value
            if x is None and headers[j0] in ('expt', 'spacer', 'design', 'ligand'):
                x = previous_row[j0]
            return x

        while not is_row_empty(i):
            row = [path] + [parse_cell(i, j, j0) for j0, j in enumerate(cols, 1)]
            rows.append(row)
            previous_row = row
            i += 1

        # Convert the parsed data to a pandas data frame.

        df = pd.DataFrame(rows, columns=headers)
        df = drop_rejected_data(df)
        df = sanitize_data(df)

        check_for_errors(df, path)

        return df
    
    except:
        msg = f"Unexpected error while parsing '{path}'"
        raise ValueError(msg)

def natsort_spacers(df):
    from natsort import order_by_index, index_natsorted
    return df.reindex(
            index=order_by_index(df.index, index_natsorted(df.spacer)),
    )

def drop_rejected_data(df):
    df = df.groupby(['expt', 'spacer', 'design']).\
            filter(lambda x: x.reject.isnull().all())
    del df['reject']
    return df

def sanitize_data(df):
    # Iron out some idiosyncracies in the data input.  For example, initially I 
    # used '+' and '−' to indicate apo and holo reactions, but later I decided 
    # that concentrations (10,000 μM) were more informative.

    df.loc[df.ligand == '−', 'ligand'] = 0      # df.loc[<row>, <col>]
    df.loc[df.ligand == '+', 'ligand'] = 10000
    df.loc[df.design == 'rxb 11', 'design'] = 'rxb 11,1'

    return df

def check_for_errors(df, path):
    # Make sure I didn't leave in any pixels I meant to set to 0.
    if (df.pixels > 100_000).any():
        msg = f"Found pixels > 100,000 in '{path}', did you forget to zero them?"
        raise ValueError(msg)

    # Make sure all the data have spacers
    if df.spacer.apply(lambda x: x.strip()).isnull().any():
        msg = f"Found data without a spacer in '{path}'"
        raise ValueError(msg)

    # Make sure none of the designs are misspelled.
    expected_designs = {'on', 'off', 'rxb 11,1', 'mhf 30', 'mhf 37'}
    unexpected_designs = set(df.design) - expected_designs
    if unexpected_designs:
        msg = f"Found the following unexpected designs in '{path}': {', '.join(str(x) for x in sorted(unexpected_designs))}"
        raise ValueError(msg)

    # Make sure none of the bands are misspelled
    expected_bands = {4000, 2000, 500, 350, 150}
    unexpected_bands = set(df.band) - expected_bands
    if unexpected_bands:
        msg = f"Found the following unexpected bands in '{path}': {', '.join(str(x) for x in sorted(unexpected_bands))}"
        raise ValueError(msg)


df = parse_xlsx_dir('densiometry')
#df = parse_xlsx('densiometry/20170916_test_d11_d12.xlsx')

#print(df)
#print(df[DISPLAY])

def percent_cut(group):
    uncut_px = group[group.band == 4000].pixels.iat[0]
    cut_px = group[group.band == 2000].pixels.iat[0]
    x = cut_px / (uncut_px + cut_px)
    return pd.Series({'percent_cut': x})

def percent_change(group):
    group = group.reset_index()  # MultiIndex confuses me...so just nuke it...
    apo_percent = group[group.ligand == 0].percent_cut.iat[0]
    holo_percent = group[group.ligand == 10000].percent_cut.iat[0]
    percent_change = holo_percent - apo_percent
    return pd.Series({'percent_change': percent_change})

def mean_change(group):
    return pd.Series({
        'mean_change': group.percent_change.mean(),
        'std_change': group.percent_change.std(),
        'num_replicates': len(group),
    })


df = df.groupby(['spacer', 'design', 'expt', 'ligand']).apply(percent_cut)
df = df.groupby(['spacer', 'design', 'expt']).apply(percent_change)
df = df.groupby(['spacer', 'design']).apply(mean_change).reset_index()
df = natsort_spacers(df)

spacers = df['spacer'].unique()
mhf = df[df.design == 'mhf 30']
rxb = df[df.design == 'rxb 11,1']

print(mhf)
print(rxb)

from color_me import ucsf
from matplotlib.colorbar import Colorbar

class MyNorm(mpl.colors.Normalize):

    def __init__(self, vmin, vmax):
        super().__init__(vmin, vmax)

    def __call__(self, value, clip=None):
        return np.ma.masked_array((value + 1) / 2)



norm = MyNorm(-0.3, 0.3)
cmap = mpl.colors.LinearSegmentedColormap.from_list('mhf_rxb', [
    (norm(-1.0), ucsf.navy[0]),
    (norm(-0.2), ucsf.navy[1]),
    (norm( 0.0), ucsf.white[0]),
    (norm( 0.2), ucsf.teal[1]),
    (norm( 1.0), ucsf.teal[0]),
])

data = np.vstack([
    mhf['mean_change'].values,
    rxb['mean_change'].values,
])

data = np.hstack((data, np.zeros((2,1))))


#fig, ax = plt.subplots(2, 2)
N = 12

from matplotlib.gridspec import GridSpec

grid = GridSpec(2,2, width_ratios=[30,1])

ax1 = plt.subplot(grid[0,0])
ax2 = plt.subplot(grid[1,0])
ax3 = plt.subplot(grid[:,1])

i=ax1.matshow(data[:,:N], cmap=cmap, norm=norm)
ax1.set_yticks([0, 1])
ax1.set_yticklabels(['mhf', 'rxb'])
ax1.set_xticks(range(N))
ax1.set_xticklabels(spacers[:N])

j=ax2.matshow(data[:,N:], cmap=cmap, norm=norm)
ax2.set_yticks([0, 1])
ax2.set_yticklabels(['mhf', 'rxb'])
ax2.set_xticks(range(N))
ax2.set_xticklabels(spacers[N:])

Colorbar(ax3, i, extend='both')
#Colorbar(ax[1,1], j)

#plt.matshow(data, cmap=cmap, norm=norm)
#plt.matshow(data, cmap=cmap, norm=norm)
#plt.colorbar(extend='both')
#plt.yticks([0, 1], ['mhf', 'rxb'])
#plt.xticks(range(len(spacers)), spacers)
fig = plt.gcf()
fig.set_size_inches(12, 5)
plt.savefig('heatmap.svg')
plt.show()
